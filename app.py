# -*- coding: utf-8 -*-
"""
Detector de Canibalizaciones SEO — La Fábrica del SEO  (v2)
===========================================================
Conecta con Ahrefs API v3 (endpoint `organic-keywords`) y detecta
canibalizaciones REALES: keywords donde 2+ URLs del proyecto rankean
simultáneamente en posiciones relevantes (no solo "Top keyword duplicada").

Para cada grupo, calcula:
  - Severidad (Alta / Media / Baja) por reglas SEO.
  - Score de impacto (volumen × CTR esperado × nº URLs) para priorizar.
  - Recomendación de Claude con contexto (volumen, KD, intención, tráfico).

Dos entradas:
  - API Ahrefs (recomendado).
  - CSV de Top Pages (fallback, formato del Colab original).
"""

from __future__ import annotations

import io
import json
import os
import time
from urllib.parse import urlparse

import pandas as pd
import streamlit as st
from anthropic import Anthropic, APIError

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Canibalizaciones SEO · La Fábrica del SEO",
    page_icon="🥩",
    layout="wide",
    initial_sidebar_state="expanded",
)

CLAUDE_MODEL = "claude-sonnet-4-5-20250929"

# CTR aproximado por posición (Sistrix / Advanced Web Ranking 2024).
# Usado para estimar tráfico potencial perdido por canibalización.
CTR_BY_POSITION = {
    1: 0.30, 2: 0.16, 3: 0.10, 4: 0.07, 5: 0.05,
    6: 0.04, 7: 0.03, 8: 0.025, 9: 0.02, 10: 0.018,
}


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------


def clean_url(url: str) -> str:
    if not isinstance(url, str):
        return ""
    o = urlparse(url.strip())
    if not o.scheme or not o.netloc:
        return url.strip()
    return f"{o.scheme}://{o.netloc}{o.path}"


def estimated_ctr(position: float) -> float:
    """CTR estimado para una posición (0 si >10)."""
    try:
        p = int(round(float(position)))
    except (TypeError, ValueError):
        return 0.0
    return CTR_BY_POSITION.get(p, 0.0)


INTENT_PRIORITY = [
    ("branded", "is_branded"),
    ("transaccional", "is_transactional"),
    ("comercial", "is_commercial"),
    ("navegacional", "is_navigational"),
    ("informacional", "is_informational"),
]


def detect_intent_for_group(group: pd.DataFrame) -> str:
    """
    Intención unificada por keyword: si CUALQUIER fila del grupo tiene un flag
    activo, lo aplicamos al grupo entero. La prioridad va de más específico
    (branded) a más genérico (informacional).
    """
    for label, col in INTENT_PRIORITY:
        if col in group.columns and (group[col] == True).any():  # noqa: E712
            return label
    return "desconocida"


def detect_page_type(url: str) -> str:
    """Heurística simple por URL: blog / categoría / producto / servicio / home."""
    u = (url or "").lower()
    if u.endswith("/") and u.count("/") <= 3:
        return "home"
    for marker, label in [
        ("/blog/", "blog"),
        ("/noticias/", "blog"),
        ("/recursos/", "blog"),
        ("/categoria/", "categoría"),
        ("/category/", "categoría"),
        ("/producto/", "producto"),
        ("/product/", "producto"),
        ("/tienda/", "producto"),
        ("/servicios/", "servicio"),
        ("/servicio/", "servicio"),
        ("/contacto", "contacto"),
        ("/sobre-", "página fija"),
        ("/about", "página fija"),
    ]:
        if marker in u:
            return label
    return "página"


# Subdirectorios típicos de idioma (ISO + algunos comunes).
LANG_PREFIXES = {
    "/en/", "/es/", "/fr/", "/de/", "/it/", "/pt/", "/ca/", "/eu/", "/gl/",
    "/nl/", "/ru/", "/zh/", "/ja/", "/ar/", "/pl/",
}


def detect_lang_from_url(url: str) -> str:
    """Devuelve el código de idioma detectado por path, o 'default'."""
    u = (url or "").lower()
    o = urlparse(u)
    path = o.path or "/"
    for prefix in LANG_PREFIXES:
        if path.startswith(prefix):
            return prefix.strip("/")
    return "default"


def classify_pattern(group: pd.DataFrame) -> str:
    """
    Clasifica el tipo de canibalización:
      - "Idiomas distintos": URLs en directorios de idioma diferentes
        (incluye home ES vs home EN — son traducciones legítimas).
      - "Falta versión idioma": una URL en /en/ y otra es la home Y la home
        NO está en idioma traducido — falta una página dedicada.
      - "Blog vs comercial": mezcla blog + categoría/producto/servicio.
      - "Producto vs categoría": un producto y una categoría compiten.
      - "Mismo tipo": todas comparten tipo (canibalización clásica).
      - "Mixto": no encaja en los anteriores.
    """
    langs = group["url"].apply(detect_lang_from_url).tolist()
    types = group["page_type"].unique() if "page_type" in group.columns else []
    type_set = set(types)

    distinct_langs = set(langs)
    explicit_langs = distinct_langs - {"default"}

    # Caso: 2+ idiomas explícitos diferentes → siempre "Idiomas distintos"
    if len(explicit_langs) >= 2:
        return "Idiomas distintos"

    # Caso especial: home + URL con prefijo de idioma
    # → SOLO es "Falta versión idioma" si la home es la única URL sin prefijo
    #   Y hay otra URL que NO es home con prefijo de idioma.
    #   Si ambas son home (home/ + /en/), es traducción legítima.
    if len(explicit_langs) == 1 and "default" in distinct_langs:
        # ¿Hay URLs que sean home y otras que no? Si todas las que tienen
        # idioma explícito son home, entonces es "Idiomas distintos" (homes traducidas).
        # Si la URL con idioma NO es home, falta versión.
        page_types = group["page_type"].tolist() if "page_type" in group.columns else [None] * len(group)
        urls_by_lang = list(zip(group["url"].tolist(), langs, page_types))
        non_default_types = {pt for u, lg, pt in urls_by_lang if lg != "default"}
        default_types = {pt for u, lg, pt in urls_by_lang if lg == "default"}

        # Si todas las URLs (default y no default) son home → idiomas distintos
        if non_default_types == {"home"} and default_types == {"home"}:
            return "Idiomas distintos"
        # Si la default es home pero la otra no → falta versión idioma
        if "home" in default_types and "home" not in non_default_types:
            return "Falta versión idioma"
        # Si la no-default es home pero la default no → falta versión idioma (al revés)
        if "home" in non_default_types and "home" not in default_types:
            return "Falta versión idioma"
        # Otros mixtos con un solo idioma explícito → idiomas distintos
        return "Idiomas distintos"

    if "blog" in type_set and (type_set & {"categoría", "producto", "servicio"}):
        return "Blog vs comercial"
    if {"producto", "categoría"}.issubset(type_set):
        return "Producto vs categoría"
    if len(type_set) == 1:
        return f"Mismo tipo ({list(type_set)[0]})"
    return "Mixto"


# ---------------------------------------------------------------------------
# Cliente Ahrefs API v3 — organic-keywords
# ---------------------------------------------------------------------------


def read_top_pages_csv(file_bytes: bytes) -> pd.DataFrame:
    """Fallback: lee CSV de Top Pages con formato del Colab original."""
    for encoding in ("utf-8", "utf-8-sig", "utf-16"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception:
            continue
    else:
        df = pd.read_csv(io.BytesIO(file_bytes))

    # Limpia BOM y espacios en nombres de columna.
    df.columns = df.columns.str.strip().str.lstrip("\ufeff")

    # Mapeo Organic Keywords export de Ahrefs → formato interno.
    rename_map = {
        "Keyword": "keyword",
        "Current URL": "url",
        "Current position": "best_position",
        "Volume": "volume",
        "KD": "keyword_difficulty",
        "Current organic traffic": "traffic",
        "Branded": "is_branded",
        "Commercial": "is_commercial",
        "Informational": "is_informational",
        "Navigational": "is_navigational",
        "Transactional": "is_transactional",
        # Fallback Top Pages legacy
        "URL": "url",
        "Top keyword": "keyword",
        "Top keyword: Position": "best_position",
        "Current traffic": "traffic",
        "Top keyword KD": "keyword_difficulty",
    }
    for k, v in rename_map.items():
        if k in df.columns and v not in df.columns:
            df = df.rename(columns={k: v})

    # Normalizar flags booleanos ("true"/"false" como string).
    for flag in ("is_branded", "is_commercial", "is_informational", "is_navigational", "is_transactional"):
        if flag in df.columns:
            df[flag] = df[flag].map({"true": True, "false": False, True: True, False: False})

    return df


# ---------------------------------------------------------------------------
# Detección + scoring
# ---------------------------------------------------------------------------


def detect_canibalizations(
    df: pd.DataFrame,
    max_position: int = 20,
    min_volume: int = 10,
) -> pd.DataFrame:
    """
    Filtra parejas (keyword, URL) que cumplen umbrales y agrupa por keyword
    canibalizada (>=2 URLs distintas).

    Espera columnas: url, keyword, best_position, volume.
    """
    required = {"url", "keyword", "best_position"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"Faltan columnas requeridas: {', '.join(missing)}. "
            "Asegúrate de usar el endpoint organic-keywords o el CSV correcto."
        )

    work = df.copy()
    work["url"] = work["url"].apply(clean_url)
    work = work.dropna(subset=["url", "keyword", "best_position"])
    work = work[work["url"] != ""]

    # Filtros: posición y volumen mínimos.
    work["best_position"] = pd.to_numeric(work["best_position"], errors="coerce")
    work = work.dropna(subset=["best_position"])
    work = work[work["best_position"] <= max_position]

    if "volume" in work.columns:
        work["volume"] = pd.to_numeric(work["volume"], errors="coerce").fillna(0)
        work = work[work["volume"] >= min_volume]

    # Canibalización = misma keyword en ≥2 URLs distintas.
    grouped = work.groupby("keyword")["url"].nunique()
    cannibal_kws = grouped[grouped > 1].index
    cannibal = work[work["keyword"].isin(cannibal_kws)].copy()

    if cannibal.empty:
        return cannibal

    # Enriquecimiento: tipo de página (por fila).
    cannibal["page_type"] = cannibal["url"].apply(detect_page_type)

    # Intención y patrón: a nivel de grupo (no por fila).
    has_intent_flags = any(c.startswith("is_") for c in cannibal.columns)
    intents_by_kw = {}
    patterns_by_kw = {}
    for kw, group in cannibal.groupby("keyword"):
        intents_by_kw[kw] = (
            detect_intent_for_group(group) if has_intent_flags else "desconocida"
        )
        patterns_by_kw[kw] = classify_pattern(group)
    cannibal["intent"] = cannibal["keyword"].map(intents_by_kw)
    cannibal["pattern"] = cannibal["keyword"].map(patterns_by_kw)

    return cannibal.sort_values(["keyword", "best_position"])


def score_group(group: pd.DataFrame) -> dict:
    """
    Calcula severidad + score de impacto para un grupo canibalizado.

    Severidad:
      - Alta: ≥2 URLs en top 10 con vol ≥100, o
              ≥1 URL en top 10 con intención comercial/transaccional y vol ≥150.
      - Media: ≥1 URL en top 10 + otra entre 11-20, o
               vol entre 50-150, o
               intención comercial/transaccional con vol ≥50.
      - Baja: branded, posiciones distantes, o vol <50.

    Score = volumen × Σ CTR(posición) × factor por nº URLs (penaliza fragmentación).
    """
    n_urls = group["url"].nunique()
    volume = float(group["volume"].iloc[0]) if "volume" in group.columns else 0.0
    positions = group["best_position"].astype(float).tolist()
    in_top10 = sum(1 for p in positions if p <= 10)
    intent = group["intent"].iloc[0] if "intent" in group.columns else "desconocida"
    is_commercial_intent = intent in {"comercial", "transaccional"}

    # Severidad
    if intent == "branded":
        severity = "Baja"
    elif in_top10 >= 2 and volume >= 100:
        severity = "Alta"
    elif in_top10 >= 1 and is_commercial_intent and volume >= 150:
        severity = "Alta"
    elif in_top10 >= 1 and any(11 <= p <= 20 for p in positions):
        severity = "Media"
    elif is_commercial_intent and volume >= 50:
        severity = "Media"
    elif 50 <= volume < 150:
        severity = "Media"
    else:
        severity = "Baja"

    # Score de impacto (tráfico estimado fragmentado)
    ctr_sum = sum(estimated_ctr(p) for p in positions)
    score = round(volume * ctr_sum * (1 + 0.1 * (n_urls - 2)), 1)

    return {"severity": severity, "score": score, "n_urls": n_urls, "in_top10": in_top10}


# ---------------------------------------------------------------------------
# Claude — prompt enriquecido
# ---------------------------------------------------------------------------

ACTION_SYSTEM_PROMPT = """Eres un consultor SEO senior español especializado en \
resolver canibalizaciones de keywords. Recibes un grupo de URLs que compiten \
por la misma keyword, con datos de Ahrefs (volumen, KD, posición, tráfico, \
intención, tipo de página, patrón de canibalización).

Devuelve UNA acción entre estas (usa el texto exacto):
- "Consolidar": fusionar URLs en una y redirigir 301 las demás (misma intención, contenido duplicado).
- "Redirigir 301": una URL es claramente superior; las demás desaparecen y se redirigen a ella.
- "Desindexar": URLs sin tráfico ni intención propia (añadir noindex o eliminar página).
- "Diferenciar intención": las URLs cubren intenciones distintas; reescribir y reorientar enlazado interno para que cada una satisfaga una intención única.
- "Reescribir y reforzar": hay una URL principal clara pero está mal optimizada o carece de autoridad; potenciar esa URL sin eliminar las demás.
- "Revisar hreflang": el patrón indica URLs en distintos idiomas o regiones; el problema es de hreflang, no de canibalización real. Corregir etiquetas antes de actuar.
- "Crear contenido": falta una URL con la intención correcta; las existentes no son el tipo idóneo para la keyword (ej. solo hay un blog post para "comprar X").
- "Mantener y monitorizar": canibalización leve (branded, posiciones distantes, volumen mínimo) sin acción urgente.

Prioriza como URL principal la que tenga mejor combinación de: posición más alta, \
mayor tráfico y coherencia con la intención de la keyword \
(ej. para "comprar X", producto o categoría > blog post).

Responde SOLO un JSON válido (sin markdown, sin texto extra):

{
  "accion": "<acción exacta de la lista>",
  "url_principal": "<URL canónica recomendada o '' si no aplica>",
  "diagnostico": "<1 frase que describe el problema raíz>",
  "accion_concreta": "<pasos específicos a ejecutar, citando URLs y datos>",
  "resultado_esperado": "<qué mejora SEO se espera tras aplicar la acción>"
}

No inventes datos que no estén en el grupo. Sé específico y cita posiciones, \
volúmenes o tipos de página cuando sea relevante."""


def build_user_prompt(keyword: str, group: pd.DataFrame, score_data: dict) -> str:
    volume = group["volume"].iloc[0] if "volume" in group.columns else "N/D"
    intent = group["intent"].iloc[0] if "intent" in group.columns else "desconocida"
    pattern = group["pattern"].iloc[0] if "pattern" in group.columns else "—"

    # KD: solo lo añadimos si Ahrefs lo devolvió no-nulo y no es 0.
    kd_str = "N/D"
    if "keyword_difficulty" in group.columns:
        kd_clean = pd.to_numeric(group["keyword_difficulty"], errors="coerce").dropna()
        if not kd_clean.empty:
            max_kd = int(kd_clean.max())
            kd_str = str(max_kd) if max_kd > 0 else "N/D"

    rows = []
    for _, r in group.iterrows():
        parts = [f"URL: {r['url']}"]
        parts.append(f"Posición: {r['best_position']}")
        if "traffic" in group.columns and pd.notna(r.get("traffic")):
            parts.append(f"Tráfico: {r['traffic']}")
        if "page_type" in group.columns:
            parts.append(f"Tipo: {r['page_type']}")
        rows.append("- " + " | ".join(parts))

    pattern_hint = ""
    if pattern == "Idiomas distintos":
        pattern_hint = (
            "\n\nNOTA: las URLs están en directorios de idioma diferentes. "
            "Esto suele NO ser canibalización real sino un problema de hreflang. "
            "Usa la acción 'Revisar hreflang' salvo que compruebes que un idioma "
            "rankea activamente para keywords del otro y le roba tráfico."
        )
    elif pattern == "Falta versión idioma":
        pattern_hint = (
            "\n\nNOTA: la home compite con una URL en un directorio de idioma. "
            "Probablemente falta la versión hreflang correcta para ese idioma. "
            "Usa la acción 'Revisar hreflang' y valida que la home tenga las "
            "etiquetas hreflang apuntando a la versión de idioma correcta."
        )

    return (
        f'Keyword canibalizada: "{keyword}"\n'
        f'Volumen: {volume} | KD: {kd_str} | Intención: {intent} | '
        f'Patrón: {pattern} | Severidad: {score_data["severity"]} | '
        f'URLs en top 10: {score_data["in_top10"]} de {score_data["n_urls"]}\n\n'
        f"URLs que compiten por esta keyword:\n"
        + "\n".join(rows)
        + pattern_hint
        + "\n\nDevuélveme la acción recomendada en el JSON descrito."
    )


def ask_claude(client: Anthropic, keyword: str, group: pd.DataFrame, score_data: dict) -> dict:
    try:
        msg = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=500,
            system=ACTION_SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": build_user_prompt(keyword, group, score_data)}
            ],
        )
        text = "".join(
            b.text for b in msg.content if getattr(b, "type", "") == "text"
        ).strip()
        if text.startswith("```"):
            text = text.strip("`")
            if text.lower().startswith("json"):
                text = text[4:].strip()
        data = json.loads(text)
        # Componer justificación: unir las 3 partes estructuradas si vienen,
        # o usar el campo justificacion directamente si el modelo lo devuelve así.
        diagnostico = data.get("diagnostico", "").strip()
        accion_concreta = data.get("accion_concreta", "").strip()
        resultado = data.get("resultado_esperado", "").strip()
        partes = [p for p in [diagnostico, accion_concreta, resultado] if p]
        justif = " | ".join(partes) if partes else data.get("justificacion", "").strip()
        return {
            "accion": data.get("accion", "").strip(),
            "url_principal": data.get("url_principal", "").strip(),
            "justificacion": justif or "Sin justificación.",
        }
    except (json.JSONDecodeError, APIError, Exception) as e:
        return {
            "accion": "Error",
            "url_principal": "",
            "justificacion": f"No se pudo generar recomendación ({type(e).__name__}).",
        }


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def build_excel(cannibal: pd.DataFrame, summary: pd.DataFrame) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()

    # Construimos el workbook manualmente para control fino sobre la hoja Diagnóstico
    wb = Workbook()
    wb.remove(wb.active)

    # ─── Hoja 1: Diagnóstico (resumen ejecutivo) ──────────────────────
    ws_diag = wb.create_sheet("Diagnóstico")

    # Estilos reutilizables
    title_font = Font(bold=True, size=16, color="1D3557")
    section_font = Font(bold=True, size=12, color="E63946")
    label_font = Font(bold=True, size=11)
    body_font = Font(size=11)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Cabecera
    ws_diag["A1"] = "Diagnóstico de Canibalizaciones SEO"
    ws_diag["A1"].font = title_font
    ws_diag.merge_cells("A1:E1")
    row = 3

    # Bloque 1: totales
    ws_diag.cell(row=row, column=1, value="RESUMEN").font = section_font
    row += 1
    ws_diag.cell(row=row, column=1, value="Total de canibalizaciones").font = label_font
    ws_diag.cell(row=row, column=2, value=len(summary))
    row += 1
    sev_counts = summary["Severidad"].value_counts()
    for sev, color in [("Alta", "F8B4B4"), ("Media", "FDE68A"), ("Baja", "BBF7D0")]:
        ws_diag.cell(row=row, column=1, value=f"  Severidad {sev}").font = label_font
        c = ws_diag.cell(row=row, column=2, value=int(sev_counts.get(sev, 0)))
        c.fill = PatternFill("solid", fgColor=color)
        row += 1
    row += 1

    # Bloque 2: distribución por patrón
    ws_diag.cell(row=row, column=1, value="DISTRIBUCIÓN POR PATRÓN").font = section_font
    row += 1
    pattern_counts = summary["Patrón"].value_counts()
    for pattern, count in pattern_counts.items():
        pct = count / len(summary) * 100
        ws_diag.cell(row=row, column=1, value=f"  {pattern}").font = body_font
        ws_diag.cell(row=row, column=2, value=f"{count} ({pct:.0f}%)").font = body_font
        row += 1
    row += 1

    # Bloque 3: conclusión accionable si hay un patrón dominante
    if len(pattern_counts) > 0:
        top_pattern = pattern_counts.index[0]
        top_count = pattern_counts.iloc[0]
        top_pct = top_count / len(summary) * 100
        if top_pct >= 40:
            ws_diag.cell(row=row, column=1, value="CONCLUSIÓN ACCIONABLE").font = section_font
            row += 1
            conclusion_map = {
                "Idiomas distintos": (
                    f"El {top_pct:.0f}% de las canibalizaciones se deben a directorios "
                    f"de idioma compitiendo entre sí. Probable causa raíz: "
                    f"implementación incorrecta de hreflang.\n\n"
                    f"PRIORIDAD: revisar hreflang en todo el sitio antes de actuar "
                    f"caso por caso. Una sola corrección técnica resuelve {top_count} "
                    f"de las {len(summary)} canibalizaciones detectadas."
                ),
                "Falta versión idioma": (
                    f"El {top_pct:.0f}% de los casos son por falta de contenido en el "
                    f"idioma de la búsqueda.\n\n"
                    f"PRIORIDAD: crear las páginas dedicadas que faltan. La estructura "
                    f"de la web tiene huecos en el idioma de las búsquedas detectadas."
                ),
                "Blog vs comercial": (
                    f"El {top_pct:.0f}% son blog canibalizando páginas comerciales.\n\n"
                    f"PRIORIDAD: diferenciar intenciones y corregir enlazado interno "
                    f"para que cada URL atraiga su tipo de búsqueda."
                ),
                "Producto vs categoría": (
                    f"El {top_pct:.0f}% son producto vs categoría.\n\n"
                    f"PRIORIDAD: definir cuál es la URL principal para cada keyword "
                    f"y redirigir o canonicalizar."
                ),
            }
            conclusion = conclusion_map.get(
                top_pattern,
                f"El patrón dominante es «{top_pattern}» ({top_pct:.0f}%)."
            )
            cell = ws_diag.cell(row=row, column=1, value=conclusion)
            cell.font = body_font
            cell.alignment = wrap_align
            ws_diag.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws_diag.row_dimensions[row].height = 80
            row += 2

    # Bloque 4: top 3 acciones críticas (mayor score de impacto)
    ws_diag.cell(row=row, column=1, value="TOP 3 ACCIONES CRÍTICAS (mayor score de impacto)").font = section_font
    row += 1
    headers = ["#", "Keyword", "Patrón", "Acción", "Score"]
    for i, h in enumerate(headers, 1):
        c = ws_diag.cell(row=row, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1D3557")
        c.font = Font(bold=True, color="FFFFFF")
    row += 1
    for i, (_, r) in enumerate(summary.head(3).iterrows(), 1):
        ws_diag.cell(row=row, column=1, value=i)
        ws_diag.cell(row=row, column=2, value=r["keyword"])
        ws_diag.cell(row=row, column=3, value=r["Patrón"])
        ws_diag.cell(row=row, column=4, value=r.get("Acción", ""))
        ws_diag.cell(row=row, column=5, value=r.get("Score impacto", ""))
        row += 1

    # Anchuras hoja Diagnóstico
    ws_diag.column_dimensions["A"].width = 50
    ws_diag.column_dimensions["B"].width = 30
    ws_diag.column_dimensions["C"].width = 25
    ws_diag.column_dimensions["D"].width = 22
    ws_diag.column_dimensions["E"].width = 12

    # ─── Hoja 2: Resumen ────────────────────────────────────────────
    ws_resumen = wb.create_sheet("Resumen")
    if "Estado" not in summary.columns:
        summary = summary.copy()
        summary["Estado"] = "Pendiente"

    for col_idx, col_name in enumerate(summary.columns, 1):
        ws_resumen.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)
    for row_idx, (_, r) in enumerate(summary.iterrows(), 2):
        for col_idx, col_name in enumerate(summary.columns, 1):
            val = r[col_name]
            if pd.isna(val):
                val = ""
            ws_resumen.cell(row=row_idx, column=col_idx, value=val)

    # Coloreado de severidad
    sev_colors = {"Alta": "F8B4B4", "Media": "FDE68A", "Baja": "BBF7D0"}
    sev_col = None
    for idx, cell in enumerate(ws_resumen[1], start=1):
        if cell.value == "Severidad":
            sev_col = idx
            break
    if sev_col:
        for row_cells in ws_resumen.iter_rows(min_row=2, min_col=sev_col, max_col=sev_col):
            cell = row_cells[0]
            fill = sev_colors.get(cell.value)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)

    # Hipervínculos en URL principal (Resumen)
    url_col_res = None
    for idx, cell in enumerate(ws_resumen[1], start=1):
        if cell.value == "URL principal":
            url_col_res = idx
            break
    if url_col_res:
        link_font = Font(color="0563C1", underline="single")
        for row_cells in ws_resumen.iter_rows(min_row=2, min_col=url_col_res, max_col=url_col_res):
            cell = row_cells[0]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = link_font

    # Auto-anchura Resumen
    for col in ws_resumen.columns:
        col_letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws_resumen.column_dimensions[col_letter].width = min(max(length + 2, 12), 60)

    # ─── Hoja 3: Detalle ────────────────────────────────────────────
    ws_detalle = wb.create_sheet("Detalle")
    detail_cols = [
        "keyword", "url", "best_position", "page_type", "intent", "pattern",
    ]
    for opt in ("volume", "keyword_difficulty", "traffic"):
        if opt in cannibal.columns:
            detail_cols.append(opt)
    detail_cols = [c for c in detail_cols if c in cannibal.columns]
    merged = cannibal[detail_cols].merge(
        summary[["keyword", "Severidad", "Score impacto", "Acción",
                 "URL principal", "Justificación"]],
        on="keyword",
        how="left",
    )
    merged["Estado"] = "Pendiente"

    for col_idx, col_name in enumerate(merged.columns, 1):
        ws_detalle.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)
    for row_idx, (_, r) in enumerate(merged.iterrows(), 2):
        for col_idx, col_name in enumerate(merged.columns, 1):
            val = r[col_name]
            if pd.isna(val):
                val = ""
            ws_detalle.cell(row=row_idx, column=col_idx, value=val)

    # Hipervínculos en url (Detalle)
    url_col_det = None
    for idx, cell in enumerate(ws_detalle[1], start=1):
        if cell.value == "url":
            url_col_det = idx
            break
    if url_col_det:
        link_font = Font(color="0563C1", underline="single")
        for row_cells in ws_detalle.iter_rows(min_row=2, min_col=url_col_det, max_col=url_col_det):
            cell = row_cells[0]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = link_font

    # Auto-anchura Detalle
    for col in ws_detalle.columns:
        col_letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws_detalle.column_dimensions[col_letter].width = min(max(length + 2, 12), 60)

    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

CUSTOM_CSS = """
<style>
    :root {
        --lfs-primary: #E63946;
        --lfs-dark: #1D3557;
        --lfs-light: #F1FAEE;
        --lfs-accent: #457B9D;
    }
    .main .block-container { padding-top: 2rem; max-width: 1300px; }
    h1 { color: var(--lfs-dark); font-weight: 700; }
    h2, h3 { color: var(--lfs-dark); }
    .stButton>button {
        background: var(--lfs-primary); color: white; border: none;
        font-weight: 600; border-radius: 6px; padding: 0.5rem 1.2rem;
    }
    .stButton>button:hover { background: var(--lfs-dark); color: white; }
    .stDownloadButton>button {
        background: var(--lfs-dark); color: white; border: none;
        font-weight: 600; border-radius: 6px;
    }
    .stDownloadButton>button:hover { background: var(--lfs-accent); color: white; }
    [data-testid="stMetricValue"] { color: var(--lfs-primary); font-weight: 700; }
    [data-testid="stSidebar"] { background-color: #FAFAFA; }
    .severity-Alta { background: #F8B4B4; padding: 2px 8px; border-radius: 4px; }
    .severity-Media { background: #FDE68A; padding: 2px 8px; border-radius: 4px; }
    .severity-Baja { background: #BBF7D0; padding: 2px 8px; border-radius: 4px; }
</style>
"""


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------


def _get_secret(key: str) -> str:
    try:
        if hasattr(st, "secrets") and key in st.secrets:
            return st.secrets[key] or ""
    except Exception:
        pass
    return os.environ.get(key, "")


def run_analysis(
    df: pd.DataFrame,
    client_name: str,
    use_claude: bool,
    max_groups: int,
    anthropic_key: str,
    max_position: int,
    min_volume: int,
    button_key: str,
) -> None:
    try:
        cannibal = detect_canibalizations(df, max_position=max_position, min_volume=min_volume)
    except ValueError as e:
        st.error(str(e))
        with st.expander("Columnas detectadas"):
            st.write(list(df.columns))
        return

    if cannibal.empty:
        st.success("✅ No se detectaron canibalizaciones con los umbrales actuales.")
        return

    # Calcular score y severidad por grupo.
    summary_records = []
    score_by_kw = {}
    for kw, group in cannibal.groupby("keyword"):
        s = score_group(group)
        score_by_kw[kw] = s

        # KD: solo lo incluimos si Ahrefs lo devuelve no-nulo en alguna fila.
        # La API marca KD como nullable y a menudo viene a None.
        kd_val = None
        if "keyword_difficulty" in group.columns:
            kd_series = pd.to_numeric(group["keyword_difficulty"], errors="coerce")
            kd_series = kd_series.dropna()
            if not kd_series.empty:
                max_kd = int(kd_series.max())
                kd_val = max_kd if max_kd > 0 else None

        summary_records.append({
            "keyword": kw,
            "Patrón": group["pattern"].iloc[0] if "pattern" in group.columns else "—",
            "Nº URLs": s["n_urls"],
            "URLs en top 10": s["in_top10"],
            "Volumen": int(group["volume"].iloc[0]) if "volume" in group.columns else None,
            "KD": kd_val,
            "Intención": group["intent"].iloc[0] if "intent" in group.columns else "desconocida",
            "Severidad": s["severity"],
            "Score impacto": s["score"],
            "Acción": "",
            "URL principal": "",
            "Justificación": "",
            "Estado": "Pendiente",
        })
    summary = pd.DataFrame(summary_records).sort_values(
        by=["Severidad", "Score impacto"],
        key=lambda c: c.map({"Alta": 0, "Media": 1, "Baja": 2}) if c.name == "Severidad" else c,
        ascending=[True, False],
    ).reset_index(drop=True)

    # Métricas de cabecera
    sev_counts = summary["Severidad"].value_counts()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Canibalizaciones", len(summary))
    c2.metric("🔴 Severidad alta", int(sev_counts.get("Alta", 0)))
    c3.metric("🟡 Severidad media", int(sev_counts.get("Media", 0)))
    c4.metric("🟢 Severidad baja", int(sev_counts.get("Baja", 0)))

    # NUEVO: resumen ejecutivo de patrones (mejora 4)
    pattern_counts = summary["Patrón"].value_counts()
    if len(pattern_counts) > 0 and pattern_counts.iloc[0] >= 2:
        top_pattern = pattern_counts.index[0]
        top_count = pattern_counts.iloc[0]
        pct = top_count / len(summary) * 100
        if pct >= 40:
            st.info(
                f"💡 **Patrón dominante detectado:** {pct:.0f}% de las canibalizaciones "
                f"({top_count} de {len(summary)}) son del tipo **«{top_pattern}»**. "
                f"Atacar la causa raíz puede resolver muchas a la vez."
            )

    st.divider()

    # Filtros sobre el resumen
    st.subheader("Resumen priorizado")
    fc1, fc2, fc3 = st.columns([1, 1, 1])
    with fc1:
        sev_filter = st.multiselect(
            "Filtrar por severidad",
            ["Alta", "Media", "Baja"],
            default=["Alta", "Media", "Baja"],
            key=f"sev_{button_key}",
        )
    with fc2:
        intent_options = sorted(summary["Intención"].dropna().unique().tolist())
        intent_filter = st.multiselect(
            "Filtrar por intención",
            intent_options,
            default=intent_options,
            key=f"int_{button_key}",
        )
    with fc3:
        pattern_options = sorted(summary["Patrón"].dropna().unique().tolist())
        pattern_filter = st.multiselect(
            "Filtrar por patrón",
            pattern_options,
            default=pattern_options,
            key=f"pat_{button_key}",
        )

    view = summary[
        summary["Severidad"].isin(sev_filter) &
        summary["Intención"].isin(intent_filter) &
        summary["Patrón"].isin(pattern_filter)
    ].reset_index(drop=True)
    st.dataframe(view, use_container_width=True, hide_index=True)

    with st.expander(f"Ver detalle ({len(cannibal)} filas URL × keyword)"):
        st.dataframe(cannibal, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("Generar informe con acciones SEO")

    if not st.button("🚀 Procesar con Claude y generar Excel", type="primary", key=button_key):
        return

    # Solo procesamos los grupos visibles tras filtros.
    keywords_to_process = view["keyword"].tolist()
    actions_records = []

    if use_claude:
        if not anthropic_key:
            st.error("Falta la Anthropic API Key. Añádela en la barra lateral.")
            return
        try:
            client = Anthropic(api_key=anthropic_key)
        except Exception as e:
            st.error(f"No se pudo inicializar Anthropic: {e}")
            return

        progress = st.progress(0.0, text="Analizando grupos…")
        total = min(len(keywords_to_process), max_groups)

        for i, kw in enumerate(keywords_to_process):
            if i >= max_groups:
                actions_records.append({
                    "keyword": kw,
                    "Acción": "Pendiente (límite alcanzado)",
                    "URL principal": "",
                    "Justificación": "",
                })
                continue

            group = cannibal[cannibal["keyword"] == kw]
            res = ask_claude(client, kw, group, score_by_kw[kw])
            actions_records.append({
                "keyword": kw,
                "Acción": res["accion"],
                "URL principal": res["url_principal"],
                "Justificación": res["justificacion"],
            })
            progress.progress((i + 1) / total, text=f"Analizando… {i + 1}/{total}")
            time.sleep(0.2)
        progress.empty()
    else:
        for kw in keywords_to_process:
            actions_records.append({
                "keyword": kw,
                "Acción": "", "URL principal": "", "Justificación": "",
            })

    actions_df = pd.DataFrame(actions_records)
    final_summary = view.drop(columns=["Acción", "URL principal", "Justificación"]).merge(
        actions_df, on="keyword", how="left"
    )

    st.subheader("Acciones recomendadas")
    st.dataframe(final_summary, use_container_width=True, hide_index=True)

    # Filtramos cannibal a solo las keywords procesadas para que el detalle case.
    cannibal_filtered = cannibal[cannibal["keyword"].isin(keywords_to_process)]
    excel_bytes = build_excel(cannibal_filtered, final_summary)
    safe_name = (client_name or "cliente").strip().replace(" ", "_")
    st.download_button(
        label="⬇️ Descargar Excel",
        data=excel_bytes,
        file_name=f"Canibalizaciones_{safe_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        key=f"dl_{button_key}",
    )


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------


def main() -> None:
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    st.title("🥩 Detector de Canibalizaciones SEO")
    st.caption(
        "Detecta canibalizaciones reales en proyectos SEO. "
        "Calcula severidad, prioriza por impacto y sugiere acciones con Claude."
    )

    # --- Sidebar -----------------------------------------------------------
    with st.sidebar:
        st.subheader("⚙️ Configuración")

        anthropic_key = st.text_input(
            "Anthropic API Key", type="password",
            value=_get_secret("ANTHROPIC_API_KEY"),
            help="En Streamlit Cloud → Secrets como ANTHROPIC_API_KEY.",
        )
        client_name = st.text_input(
            "Nombre del cliente", placeholder="Ej. Cronomía",
            help="Se usa en el nombre del Excel.",
        )

        st.divider()
        st.markdown("**Umbrales de detección**")
        max_position = st.slider(
            "Posición máxima a considerar", 5, 50, 20,
            help="Solo se cuentan keywords donde una URL rankea por encima de esta posición.",
        )
        min_volume = st.number_input(
            "Volumen mínimo de búsqueda", 0, 10000, 10, step=10,
            help="Filtra keywords con volumen menor (cola larga residual).",
        )

        st.divider()
        use_claude = st.toggle(
            "Generar recomendaciones con Claude", value=True,
        )
        max_groups = st.number_input(
            "Máx. grupos a analizar con Claude",
            min_value=1, max_value=500, value=50, step=10,
        )

    # --- CSV ---------------------------------------------------------------
    st.markdown(
        "**Cómo exportar el CSV:** entra en Ahrefs → Site Explorer → introduce el dominio del cliente "
        "→ menú izquierdo **Organic keywords** → botón **Exportar** (arriba a la derecha)."
    )
    uploaded = st.file_uploader(
        "Sube el CSV de Organic Keywords", type=["csv"], accept_multiple_files=False,
    )
    if uploaded:
        try:
            df_csv = read_top_pages_csv(uploaded.getvalue())
        except Exception as e:
            st.error(f"No se pudo leer el CSV: {e}")
            return
        st.success(f"CSV cargado: {len(df_csv):,} filas.")
        st.divider()
        run_analysis(
            df_csv, client_name, use_claude, int(max_groups),
            anthropic_key, int(max_position), int(min_volume),
            button_key="btn_csv",
        )


if __name__ == "__main__":
    main()
